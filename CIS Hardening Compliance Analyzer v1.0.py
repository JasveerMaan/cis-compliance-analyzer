# =============================================================================
# CIS Hardening Compliance Analyzer
# =============================================================================
# Compares Nessus compliance scan CSVs against CIS hardening guide Excel
# templates and produces two Excel reports:
#   1. CIS Compliance Report.xlsx  — per-host control results
#   2. CIS Review.xlsx             — missing controls + newer CIS version FYI
# =============================================================================

import csv, re, os, sys, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter

# ── Optional colour support in terminal output ────────────────────────────────
try:
    import colorama; colorama.init(autoreset=True); HAS_COLOR = True
except ImportError:
    HAS_COLOR = False  # Falls back to plain text if colorama is not installed

# ── Terminal colour helpers ───────────────────────────────────────────────────
# These wrap text in ANSI escape codes for coloured terminal output.
# If colorama is not installed, they return plain text instead.
def _c(t, c): return "\033[" + c + "m" + str(t) + "\033[0m" if HAS_COLOR else str(t)
def cyan(t):   return _c(t, "96")
def green(t):  return _c(t, "92")
def yellow(t): return _c(t, "93")
def red(t):    return _c(t, "91")
def bold(t):   return _c(t, "1")
def dim(t):    return _c(t, "2")
def blue(t):   return _c(t, "94")

# ── Tool metadata ─────────────────────────────────────────────────────────────
TOOL_NAME = "CIS Hardening Compliance Analyzer"
VERSION   = "1.0"
AUTHORS   = "Jasveer  |  Chen Yung"

def print_banner():
    """Print the ASCII banner at startup."""
    L = "=" * 62
    print(); print(cyan(L))
    print(cyan("  ") + bold(cyan("  " + TOOL_NAME + "  v" + VERSION)))
    print(cyan("  ") + cyan("  Compares Nessus scan results against CIS hardening guides"))
    print(cyan("  ") + cyan("  Outputs: Compliance Report + Review workbooks"))
    print(cyan("  ") + dim("  Authors: " + AUTHORS))
    print(cyan(L)); print()

# ── Logging helpers ───────────────────────────────────────────────────────────
# These print formatted status messages to the terminal during execution.
def step(n, t): print(bold(yellow("\nSTEP " + str(n) + " -- " + t)))
def ok(t):    print(green("  [OK] " + str(t)))
def warn(t):  print(yellow("  [!!] " + str(t)))
def err(t):   print(red("  [XX] " + str(t)))
def found(t): print(cyan("    *  " + str(t)))

class ProgressBar:
    """
    Simple terminal progress bar.
    Shows percentage completion and a suffix string (e.g. rows processed).
    Usage:
        pb = ProgressBar(total=100)
        pb.update(50, "50/100 rows")
        pb.done("Complete")
    """
    def __init__(self, total, width=36):
        self.total = max(total, 1); self.width = width

    def update(self, n, suffix=""):
        pct = min(n, self.total) / self.total; filled = int(self.width * pct)
        bar = green("#" * filled) + dim("-" * (self.width - filled))
        print("\r    [" + bar + "] " + yellow(str(int(pct*100)).rjust(3)+"%") +
              "  " + str(suffix), end="", flush=True)

    def done(self, msg=""):
        print("\r    [" + green("#"*self.width) + "] " + green("100%") + "  " + str(msg))

# ── Excel cell fill colours ───────────────────────────────────────────────────
# Used to colour-code rows in the output Excel report based on result.
FILL_PASS = PatternFill("solid", fgColor="C6EFCE")   # Green  — PASSED
FILL_FAIL = PatternFill("solid", fgColor="FFC7CE")   # Red    — FAILED
FILL_WARN = PatternFill("solid", fgColor="FFEB9C")   # Yellow — WARNING / NOT IN SCAN
FILL_NO   = PatternFill("solid", fgColor="D9D9D9")   # Grey   — Not Adopted
FILL_NA   = PatternFill("solid", fgColor="EDEDED")   # Light  — N/A
FILL_HDR  = PatternFill("solid", fgColor="1F4E79")   # Dark blue — Header rows
FILL_REV  = PatternFill("solid", fgColor="FCE4D6")   # Peach  — Review items
FILL_GRAY = PatternFill("solid", fgColor="F2F2F2")   # Light grey — FYI items

# ── Excel font / alignment / border presets ───────────────────────────────────
FNT_HDR   = Font(bold=True, color="FFFFFF", size=11)
FNT_TITLE = Font(bold=True, size=13, color="1F4E79")
ALIGN_WR  = Alignment(wrap_text=True, vertical="top")
ALIGN_CTR = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin      = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_hdr(ws, rn, ncols, fill=FILL_HDR):
    """Apply header styling (dark background, white bold font, border) to a row."""
    for c in range(1, ncols+1):
        cell = ws.cell(rn, c)
        cell.fill = fill; cell.font = FNT_HDR
        cell.alignment = ALIGN_CTR; cell.border = BORDER

def set_widths(ws, widths):
    """Set column widths for a worksheet from a list of pixel widths."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def rfill(result):
    """
    Return the appropriate fill colour for a cell based on the result string.
    E.g. "PASSED" → green, "FAILED" → red, "WARNING" → yellow, etc.
    Returns None if no match (cell left unstyled).
    """
    r = (result or "").upper()
    if "PASSED"      in r: return FILL_PASS
    if "FAILED"      in r: return FILL_FAIL
    if "WARNING"     in r: return FILL_WARN
    if "NOT ADOPTED" in r: return FILL_NO
    if "N/A"         in r: return FILL_NA
    return None

# ── Input helpers ─────────────────────────────────────────────────────────────
def strip_quotes(s):
    """Remove surrounding single or double quotes from a string (e.g. drag-dropped paths)."""
    s = s.strip()
    if len(s) >= 2 and s[0] in ('"', "'") and s[-1] == s[0]: s = s[1:-1]
    return s.strip()

def prompt_path(prompt_text):
    """Prompt the user for a file/folder path and keep asking until a valid one is entered."""
    while True:
        path = strip_quotes(input(prompt_text))
        if os.path.exists(path): return path
        err("Path not found: " + path)
        warn('Check for typos, or wrap paths with spaces in double quotes.\n')

def confirm(msg):
    """Ask the user a yes/no question and return True/False. Loops until valid input."""
    while True:
        ans = input(bold("  ? ") + msg + " [y/n]: ").strip().lower()
        if ans in ("y","yes"): return True
        if ans in ("n","no"):  return False
        warn("Please enter y or n.")

# ── OS/guide type detection ───────────────────────────────────────────────────
# Colour-coded tags used when listing detected guides in the terminal.
GUIDE_TAGS = {
    "windows": green("[Windows]"),
    "rhel":    blue("[RHEL]   "),
    "esxi":    cyan("[ESXi]   "),
    "aix":     cyan("[AIX]    "),
    "cisco":   cyan("[Cisco]  "),
    "aruba":   cyan("[Aruba]  "),
    "oracle":  cyan("[Oracle] "),
    "ping":    cyan("[PING]   "),
    "rsa":     cyan("[RSA]    "),
}

def detect_guide_type(filename):
    """
    Detect the OS/platform type of a guide or scan file from its filename.
    Returns a string key matching GUIDE_TAGS (e.g. "windows", "rhel") or None.
    """
    f = filename.lower()
    if "windows" in f or re.search(r"(?<![a-z])win(?:dows)?(?![a-z])", f): return "windows"
    if re.search(r"(?<![a-z])r[eh]hl?(?![a-z])", f) or "red hat" in f or "redhat" in f: return "rhel"
    if "esxi" in f or "vmware" in f: return "esxi"
    if "aix" in f: return "aix"
    if "cisco" in f or "apic" in f or "nexus" in f or "aci" in f: return "cisco"
    if "aruba" in f: return "aruba"
    if "oracle" in f: return "oracle"
    if "ping" in f: return "ping"
    if "rsa" in f: return "rsa"
    return None

def detect_win_version(csv_path):
    """
    Detect the Windows Server version (e.g. "2019", "2022") from a Nessus scan CSV.
    Searches the Description column for the CIS audit reference string.
    Falls back to extracting a 4-digit year from the filename.
    Returns the year string or None.
    """
    try:
        with open(csv_path, encoding="utf-8", errors="replace") as f:
            for row in csv.DictReader(f):
                m = re.search(r'CIS_Microsoft_Windows(?:_Server)?_(20\d{2})', row.get("Description",""))
                if m: return m.group(1)
    except Exception: pass
    m = re.search(r'20\d{2}', os.path.basename(csv_path))
    return m.group(0) if m else None

def detect_rhel_version(csv_path):
    """
    Detect the RHEL major version (e.g. "7", "8", "9") from a Nessus scan CSV.
    Checks the filename first, then scans the Description column.
    Returns the version string or None.
    """
    fname = os.path.basename(csv_path).lower()
    m = re.search(r'rhel\s*(\d+)', fname)
    if m: return m.group(1)
    try:
        with open(csv_path, encoding="utf-8", errors="replace") as f:
            for row in csv.DictReader(f):
                m2 = re.search(r'CIS_Red_Hat_Enterprise_Linux_(\d+)', row.get("Description",""))
                if m2: return m2.group(1)
    except Exception: pass
    return None

def find_guide_for_version(ver, gtype, xlsx_files):
    """
    Return all guide Excel filenames that match both the OS type and version string.
    E.g. ver="2019", gtype="windows" matches "MS Windows Server 2019_template.xlsx"
    """
    return [f for f in xlsx_files if detect_guide_type(f) == gtype and ver in f]

# ── Title normalisation ───────────────────────────────────────────────────────
# Unicode "smart quote" → ASCII quote translation table.
# Nessus and Excel guides may use different quote characters for the same string.
_QUOTE_TABLE = str.maketrans({
    "\u2019": "'", "\u2018": "'", "\u201a": "'", "\u201b": "'",
    "\u201c": '"', "\u201d": '"', "\u201e": '"', "\u201f": '"',
    "\u2032": "'", "\u2033": '"', "`":           "'"
})

def norm_title(s):
    """
    Normalise a CIS control title for comparison between guide and scan.

    Steps applied:
    1. Strip leading level marker e.g. "(L1) " or "(L2) "
    2. Replace Unicode/smart quotes with ASCII equivalents
    3. Strip editorial parentheticals immediately before "' is set to"
       — CIS guides sometimes add notes like "(not recommended)" or
         "(protects against packet spoofing)" that Nessus omits.
       Example: "Enable Automatic Logon (not recommended)' is set to 'Disabled'"
             →  "Enable Automatic Logon' is set to 'Disabled'"
    4. Collapse whitespace and lowercase

    Returns the normalised string.
    """
    s = re.sub(r"^\(L[12]\)\s*", "", str(s or ""))
    s = s.translate(_QUOTE_TABLE)
    s = re.sub(r"(\s*\([^)]+\))+(?='\s*is\s*set\s*to)", "", s, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", s).strip().lower()

def is_cis_sn(s):
    """Return True if the string looks like a CIS S/N (e.g. "1.1.1", "17.6.3")."""
    return bool(re.match(r"^\d+(\.\d+)+$", str(s or "").strip()))

# ── Guide loader ──────────────────────────────────────────────────────────────
def load_guide_excel(xlsx_path):
    """
    Load a CIS hardening guide Excel template into a dictionary keyed by S/N.

    Expected columns (auto-detected by header name):
        S/N | Title | Adoption | ... | Justification

    Adoption values recognised: YES / NO / NA
    Rows with unrecognised adoption values are skipped.

    Returns:
        dict keyed by S/N string, each value:
            {
                "title":      raw title string,
                "norm_title": normalised title (for matching),
                "norm":       "YES" / "NO" / "NA",
                "adopt":      raw adoption cell value,
                "just":       justification text
            }
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows: return {}

    # Auto-detect column positions from the header row
    headers = [str(c or "").strip().lower() for c in all_rows[0]]
    adopt_col = just_col = None
    for i, h in enumerate(headers):
        if h == "adoption ms": adopt_col = i; break
        if "adoption" in h and adopt_col is None: adopt_col = i
    for i, h in enumerate(headers):
        if "justification" in h: just_col = i; break
    if adopt_col is None: return {}

    total = len(all_rows) - 1
    pb = ProgressBar(max(total, 1))
    items = {}; seen = set()

    for idx, row in enumerate(all_rows[1:], 1):
        if idx % 20 == 0 or idx == total:
            pb.update(idx, str(idx)+"/"+str(total)+" rows  |  "+str(len(items))+" items found")

        sn = str(row[0] or "").strip()
        if not is_cis_sn(sn) or sn in seen: continue
        seen.add(sn)

        title = str(row[1] or "").replace("\n", " ").strip()
        adopt = str(row[adopt_col] or "").strip()
        nrm   = adopt.upper().replace(".", "").replace("/", "").strip()
        if nrm not in ("YES", "NO", "NA"): continue

        just = str(row[just_col] or "").replace("\n", " ").strip() if just_col is not None else ""
        items[sn] = {"title": title, "norm_title": norm_title(title),
                     "norm": nrm, "adopt": adopt, "just": just}

    pb.done(green(str(len(items)) + " items loaded"))
    return items

# ── Scan CSV parser ───────────────────────────────────────────────────────────
def parse_desc(description):
    """
    Extract structured data from a Nessus Description field.

    Nessus compliance check descriptions follow this format:
        "17.5.1 (L1) Ensure 'Audit Account Lockout' is set to..." : [PASSED]
        ...
        Policy Value:
        'failure' || 'success, failure'

        Actual Value:
        'success, failure'

    Returns:
        (sn, title, actual_value, policy_value)
        All strings; sn/title are None if the format is not recognised.
    """
    if not description: return None, None, "", ""
    m = re.match(r'^"(\d+(?:\.\d+)+)\s+(.*?)"', description.strip(), re.DOTALL)
    if not m: return None, None, "", ""

    sn    = m.group(1)
    title = re.sub(r"^\(L[12]\)\s+", "", m.group(2).replace("\n", " ").strip())
    actual = policy = ""

    # Extract "Actual Value:" block (the real system value)
    av = re.search(r"Actual Value:\s*\n(.*?)(?:\n\n|\Z)", description, re.DOTALL)
    if av: actual = av.group(1).strip().replace("\n", " ")[:400]

    # Extract "Policy Value:" block (the expected/configured value)
    pv = re.search(r"Policy Value:\s*\n(.*?)(?:\n\n|Actual Value|\Z)", description, re.DOTALL)
    if pv: policy = pv.group(1).strip().replace("\n", " ")[:200]

    return sn, title, actual, policy

def load_scan_per_host(path):
    """
    Parse a Nessus compliance scan CSV and group results by host IP.

    Only rows with Risk = PASSED / FAILED / WARNING are loaded.
    Duplicate S/N entries per host are ignored (first occurrence wins).

    Returns:
        hosts      — dict: { host_ip: { sn: {risk, actual, policy, title} } }
        title_idx  — dict: { host_ip: { norm_title: sn } }
                     (inverted title index used for exact-title lookup)
    """
    with open(path, encoding="utf-8", errors="replace") as f:
        rows = list(csv.DictReader(f))

    total = len(rows)
    hosts = defaultdict(dict)
    title_idx = defaultdict(dict)
    pb = ProgressBar(total)

    for i, r in enumerate(rows):
        if i % 25 == 0 or i == total-1:
            pb.update(i+1, str(i+1)+"/"+str(total)+" rows  |  "+str(len(hosts))+" host(s) found")

        host = r.get("Host","").strip()
        risk = r.get("Risk","").strip()
        if not host or risk not in ("PASSED","FAILED","WARNING"): continue

        sn, title, actual, policy = parse_desc(r.get("Description",""))
        if sn and sn not in hosts[host]:
            hosts[host][sn] = {"risk":risk,"actual":actual,"policy":policy,"title":title}
            title_idx[host][norm_title(title)] = sn

    pb.done(green(str(len(hosts))+" host(s) parsed  |  "+str(total)+" rows processed"))
    return dict(hosts), dict(title_idx)

def host_has_results(d):
    """
    Return True if a host has at least one PASSED or FAILED result.
    Hosts with only WARNINGs (e.g. audit failed to run) are considered invalid.
    """
    return any(v["risk"] in ("PASSED","FAILED") for v in d.values())

# ── Fuzzy title matching ──────────────────────────────────────────────────────
# FUZZY_THRESHOLD: minimum score required to accept a fuzzy title match.
# Raised to 0.82 (from 0.80) to avoid false matches on audit policy checks
# where two different subcategories share nearly identical boilerplate titles
# (e.g. "Audit Distribution Group Management" vs "Audit Application Group Management").
FUZZY_THRESHOLD = 0.82

# Caches to avoid recomputing per-host word indexes and fuzzy results.
_host_word_index = {}   # { id(title_idx_host): { norm_title: frozenset(words) } }
_fuzzy_cache     = {}   # { (guide_norm_t, id(title_idx_host)): best_scan_sn }

def _get_word_index(title_idx_host):
    """
    Build (and cache) a word-set index for all scan titles on a given host.
    Maps each normalised scan title → frozenset of words (length > 2).
    Used to efficiently compute word-Jaccard similarity during fuzzy matching.
    """
    hid = id(title_idx_host)
    if hid not in _host_word_index:
        _host_word_index[hid] = {
            nt: frozenset(w for w in nt.split() if len(w) > 2)
            for nt in title_idx_host
        }
    return _host_word_index[hid]

def _title_jaccard(norm_a, norm_b):
    """
    Compute word-Jaccard similarity between two normalised title strings.
    Jaccard = |intersection| / |union| of word sets (words length > 2).
    Returns a float in [0, 1].
    """
    wa = frozenset(w for w in norm_a.split() if len(w) > 2)
    wb = frozenset(w for w in norm_b.split() if len(w) > 2)
    inter = len(wa & wb); union = len(wa | wb)
    return inter / union if union else 0

def _lookup_scan(sn, guide_norm_t, raw_host, title_idx_host):
    """
    Find the scan result entry that best matches a guide control.

    Three-level lookup strategy:

    Level 1 — Exact title match
        Look up the normalised guide title directly in the scan's title index.
        This handles CIS S/N renumbering between guide and Nessus audit versions
        where the title is identical but the number has shifted.
        Example: Guide 17.6.1 = Nessus 17.6.3 (same title, different S/N).

    Level 2 — Same S/N with title verification (Jaccard ≥ 0.40)
        If the same S/N exists in the scan, confirm titles are about the same
        topic using a low similarity threshold (0.40).
        This handles cases where guide and Nessus phrase the same check
        differently without changing the S/N, e.g.:
            Guide:  "...is set to 'Success and Failure'"
            Nessus: "...is set to include 'Failure'"
        Without this step, the fuzzy stage could jump to a different S/N
        that scores higher due to shared boilerplate words.

    Level 3 — Fuzzy cross-S/N match (score ≥ FUZZY_THRESHOLD)
        For remaining mismatches, compute a combined similarity score:
          - Word-Jaccard: |intersection| / |union|
          - Perfect coverage: 1.0 only when ALL guide words appear in the scan
            title — i.e. Nessus APPENDED extra words (e.g. "(MS only)", IIS
            conditions) rather than SUBSTITUTING a key word.
            Substitution keeps inter < len(gw) so coverage stays at 0,
            preventing false matches between audit subcategories.
        Uses a result cache keyed by (guide_norm_t, host_id) to avoid
        re-running the O(n) search for the same guide title.

    Level 4 — Pure S/N fallback
        Last resort: use raw_host.get(sn) directly even without title check.

    Returns:
        (scan_entry_dict, sn_used, method_string)
        All three are None if no match found at any level.
    """
    # Level 1: Exact title match
    if title_idx_host:
        alt_sn = title_idx_host.get(guide_norm_t)
        if alt_sn and alt_sn in raw_host:
            return raw_host[alt_sn], alt_sn, "exact_title"

    # Level 2: Same S/N exists in scan — verify titles are about the same topic
    if sn in raw_host:
        s = raw_host[sn]
        sim = _title_jaccard(guide_norm_t, norm_title(s["title"]))
        if sim >= 0.40:
            return s, sn, "sn_verified"

    # Level 3: Fuzzy cross-S/N match
    if title_idx_host:
        cache_key = (guide_norm_t, id(title_idx_host))
        if cache_key not in _fuzzy_cache:
            gw   = frozenset(w for w in guide_norm_t.split() if len(w) > 2)
            widx = _get_word_index(title_idx_host)
            best = 0.0; best_nt = None

            for nt, sw in widx.items():
                inter = len(gw & sw)
                if inter == 0: continue
                union   = len(gw | sw)
                jaccard = inter / union if union else 0

                # Perfect coverage: fires only when guide title is a strict
                # subset of the scan title (Nessus extended it with extra words).
                # Does NOT fire when one key word is substituted for another.
                coverage = 1.0 if (len(gw) >= 6 and inter == len(gw)) else 0

                score = max(jaccard, coverage)
                if score > best: best = score; best_nt = nt

            _fuzzy_cache[cache_key] = (
                title_idx_host[best_nt] if best >= FUZZY_THRESHOLD and best_nt else None)

        csn = _fuzzy_cache[cache_key]
        if csn and csn in raw_host:
            return raw_host[csn], csn, "fuzzy_title"

    # Level 4: Pure S/N fallback (last resort)
    s = raw_host.get(sn)
    if s: return s, sn, "sn"
    return None, None, None

# ── Report builder ────────────────────────────────────────────────────────────
# Column definitions for the per-host compliance sheet
COL_HDRS = ["S/N","Title (Hardening Guide)","Adoption","Justification",
            "Nessus Result","Expected Value","Actual Scanned Value"]
WIDTHS   = [14, 55, 12, 38, 30, 22, 42]
CTR_COLS = (1, 3, 5)  # Columns to centre-align (S/N, Adoption, Nessus Result)

def build_host_rows(guide, raw_host, title_idx_host=None):
    """
    Build the row data for one host's compliance sheet.

    For each guide control:
    - Adoption = NA  → mark N/A (not applicable for this server role)
    - Adoption = NO  → intentionally not adopted; look up Nessus result for reference
    - Adoption = YES → look up Nessus result; if not found mark NOT IN SCAN

    Returns a list of tuples:
        (sn, title, adoption, justification, result, policy_value, actual_value)
    """
    rows = []
    for sn, g in sorted(guide.items()):
        norm = g["norm"]; adopt = g["adopt"]; just = g["just"]
        title = g["title"]; gnorm_t = g["norm_title"]

        if norm == "NA":
            # Control is not applicable to this server role (e.g. DC-only check on MS)
            rows.append((sn, title, adopt, just, "N/A", "", ""))

        elif norm == "NO":
            # Intentionally not adopted — still show Nessus result for reference
            s, sn_used, _ = _lookup_scan(sn, gnorm_t, raw_host, title_idx_host)
            result = ("Not Adopted (Intentional) - Nessus: " + s["risk"]) if s \
                else "Not Adopted (Intentional)"
            rows.append((sn, title, adopt, just, result,
                         s["policy"] if s else "",
                         (s["actual"] or s["policy"]) if s else ""))

        else:  # norm == "YES" — control must be adopted
            s, sn_used, _ = _lookup_scan(sn, gnorm_t, raw_host, title_idx_host)
            if s:
                # Append a note if S/N was remapped (guide vs scan S/N differ)
                note = (" (guide "+sn+" -> scan "+sn_used+")") if sn_used != sn else ""
                rows.append((sn, title, adopt, just,
                             s["risk"]+note, s["policy"], s["actual"] or s["policy"]))
            else:
                # Control not found in scan — typically DC/role-specific checks
                # that Nessus does not audit for this server type (e.g. 17.2.x on MS)
                rows.append((sn, title, adopt, just, "NOT IN SCAN", "", ""))

    return rows

def write_host_sheet(wb, sheet_name, rows, host_ip, label):
    """
    Write one host's compliance rows to a new sheet in the workbook.
    Applies colour fills, borders, alignment and auto-filter.
    """
    ws = wb.create_sheet(sheet_name)

    # Title row
    ws.append(["Host: "+host_ip+"  |  "+label])
    ws.merge_cells("A1:"+get_column_letter(len(COL_HDRS))+"1")
    ws["A1"].font = Font(bold=True, size=11, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 18

    # Header row
    ws.append(COL_HDRS); style_hdr(ws, 2, len(COL_HDRS))

    # Data rows
    for row in rows:
        sn, title, adopt, just, result, pv, av = row
        fill = rfill(result)
        ws.append([sn, title, adopt, just, result, pv, av])
        rn = ws.max_row
        for c in range(1, len(COL_HDRS)+1):
            ws.cell(rn,c).border    = BORDER
            ws.cell(rn,c).alignment = ALIGN_CTR if c in CTR_COLS else ALIGN_WR
            if fill: ws.cell(rn,c).fill = fill

    set_widths(ws, WIDTHS)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:"+get_column_letter(len(COL_HDRS))+str(ws.max_row)

# ── Review workbook builders ──────────────────────────────────────────────────
def build_review(guide, all_valid_hosts, type_label, title_idxs=None):
    """
    Build the list of controls that require manual verification.

    A control is flagged for review if it has Adoption=YES in the guide
    but could not be found in the Nessus scan on one or more hosts.

    Returns a list of tuples (one per missing control per host group):
        (source_label, host_detail, sn, title, adoption, justification, reason, detail)
    """
    review = []
    for sn in sorted(guide):
        g = guide[sn]
        if g["norm"] != "YES": continue
        gnorm_t = g["norm_title"]

        # Determine which hosts are missing this control
        missing = [h for h in sorted(all_valid_hosts)
                   if not _lookup_scan(sn, gnorm_t, all_valid_hosts[h],
                                       title_idxs.get(h) if title_idxs else None)[1]]
        present = [h for h in sorted(all_valid_hosts) if h not in missing]
        if not missing: continue

        reason = ("Not found in Nessus scan (any host)" if not present
                  else "Not found in Nessus scan ("+str(len(missing))+"/"+
                       str(len(all_valid_hosts))+" hosts)")
        detail = ("No matching compliance check for "+sn+". May not be auditable via Nessus."
                  if not present
                  else "Missing from: "+", ".join(missing)+
                       ". Present in: "+", ".join(present)+".")
        review.append((type_label, "see detail", sn, g["title"],
                       g["adopt"], g["just"], reason, detail))
    return review

def build_fyi(all_valid_hosts, guide, type_label, title_idxs=None):
    """
    Build the FYI list of Nessus checks that have no corresponding guide entry.

    These are controls present in the Nessus audit but not in the hardening
    guide — typically because Nessus uses a newer CIS benchmark version.
    Surfaced in the "Newer CIS Version (FYI)" sheet for awareness.

    Returns a list of tuples in the same format as build_review.
    """
    seen_sn = {}
    for host in sorted(all_valid_hosts):
        tidx_h = title_idxs.get(host) if title_idxs else None

        # Build the set of scan S/Ns that were matched to a guide entry
        matched_sns = set()
        for gsn, g in guide.items():
            _, sn_used, _ = _lookup_scan(gsn, g["norm_title"], all_valid_hosts[host], tidx_h)
            if sn_used: matched_sns.add(sn_used)

        # Any scan S/N not matched to a guide entry is a "newer CIS" FYI item
        for sn, s in all_valid_hosts[host].items():
            if sn not in matched_sns:
                if sn not in seen_sn:
                    seen_sn[sn] = {"title":s["title"],"hosts":[],
                                   "policy":s["policy"],"result":s["risk"]}
                seen_sn[sn]["hosts"].append(host)

    return [(type_label+" scan-only", str(len(d["hosts"]))+" hosts",
             sn, d["title"], "", "",
             "Nessus check not in hardening guide (newer CIS version)",
             "Result sample: "+d["result"]+" | Hosts: "+", ".join(d["hosts"]))
            for sn, d in sorted(seen_sn.items())]

# ── Interactive guide assignment ──────────────────────────────────────────────
def guide_matching_step(win_entries, rhel_entries, guide_dir, xlsx_files):
    """
    Present the auto-detected guide mapping to the user for confirmation/override.

    For each scan file, the script auto-detects the OS version and proposes the
    matching guide template. The analyst can override any assignment before
    processing begins.

    Status tags shown:
        [AUTO]    — one guide matched automatically
        [MULTI]   — multiple guides found; analyst should confirm
        [NO MATCH]— no guide found for this scan
        [CHANGED] — analyst manually overrode the assignment

    Returns:
        win_map  — { csv_path: guide_filename } for Windows scans
        rhel_map — { csv_path: guide_filename } for RHEL scans
    """
    print(); print(bold(yellow("  Proposed Guide Mapping")))
    print(dim("  "+"-"*70))
    entries = []; idx = 1

    # Build entry list for Windows scans
    for e in win_entries:
        fname = os.path.basename(e["csv_path"]); ver = e["ver"]
        cands = find_guide_for_version(ver, "windows", xlsx_files) if ver else []
        entries.append({"idx":idx,"csv_path":e["csv_path"],"csv_name":fname,
                        "gtype":"windows","ver":ver,"proposed":cands[0] if cands else None,
                        "candidates":cands}); idx += 1

    # Build entry list for RHEL scans
    for e in rhel_entries:
        fname = os.path.basename(e["csv_path"]); ver = e["ver"]
        cands = find_guide_for_version(ver, "rhel", xlsx_files) if ver else []
        if not cands: cands = [f for f in xlsx_files if detect_guide_type(f) == "rhel"]
        entries.append({"idx":idx,"csv_path":e["csv_path"],"csv_name":fname,
                        "gtype":"rhel","ver":ver,"proposed":cands[0] if cands else None,
                        "candidates":cands}); idx += 1

    def _tbl(entries):
        """Print the current mapping table."""
        print()
        print(dim("  {:<4}  {:<45}  {:<9}  {:<40}  {}".format(
            "#","Scan File","OS Ver","Guide Template","Status")))
        print(dim("  "+"-"*106))
        for e in entries:
            st = (green("[AUTO]")        if e["proposed"] and len(e["candidates"])==1
                  else yellow("[MULTI]") if len(e["candidates"])>1
                  else red("[NO MATCH]")  if not e["proposed"]
                  else yellow("[CHANGED]") if e.get("manual") else green("[AUTO]"))
            gs  = e["proposed"] or "(none)"
            ver = (e["gtype"][:3].upper()+str(e["ver"])) if e["ver"] else e["gtype"][:7].upper()
            print("  {:<4}  {:<45}  {:<9}  {:<40}  {}".format(
                str(e["idx"])+".", e["csv_name"][:44], ver, gs[:39], st))
        print()

    _tbl(entries)

    # Allow analyst to override individual assignments
    while True:
        ans = input(bold("  ? ")+"Enter to confirm, or # to change a guide assignment: ").strip()
        if ans == "": break
        try:
            n = int(ans)
            tgt = next((e for e in entries if e["idx"]==n), None)
            if not tgt: warn("No entry #"+str(n)); continue
        except ValueError:
            warn("Enter a number or press Enter."); continue

        gtype = tgt["gtype"]
        avail = [f for f in xlsx_files if detect_guide_type(f) == gtype]
        if not avail: warn("No "+gtype+" guides found."); continue
        print("\n  Available "+gtype.upper()+" guides:")
        for i, f in enumerate(avail, 1): print("    "+str(i)+". "+f)
        print("    0. Skip this scan file (no guide)")
        while True:
            pick = input("  Select guide # for "+tgt["csv_name"]+": ").strip()
            try:
                pi = int(pick)
                if pi == 0: tgt["proposed"] = None; tgt["manual"] = True; break
                if 1 <= pi <= len(avail): tgt["proposed"] = avail[pi-1]; tgt["manual"] = True; break
                warn("Enter a valid number.")
            except ValueError:
                warn("Enter a number.")
        _tbl(entries)

    # Warn and confirm if any scans have no guide assigned
    unmatched = [e for e in entries if not e["proposed"]]
    if unmatched:
        print()
        for e in unmatched: warn("No guide for: "+e["csv_name"]+" -- will be SKIPPED.")
        if not confirm("Proceed anyway?"): print(yellow("  Aborted.")); sys.exit(0)

    win_map  = {e["csv_path"]:e["proposed"] for e in entries if e["gtype"]=="windows" and e["proposed"]}
    rhel_map = {e["csv_path"]:e["proposed"] for e in entries if e["gtype"]=="rhel"    and e["proposed"]}
    return win_map, rhel_map

# ── Main orchestration ────────────────────────────────────────────────────────
def main():
    """
    Main entry point — orchestrates all processing steps:

    STEP 1  Prompt for hardening guide folder (Excel templates)
    STEP 2  Prompt for Nessus scan folder (CSV files)
    STEP 3  Prompt for output folder
    STEP 4  Parse all scan CSVs, group results by host
    STEP 5  Classify each scan as Windows / RHEL / unknown
    STEP 6  Present auto-detected guide mapping for analyst confirmation
    STEP 7  Load guide Excel templates (cached to avoid reloading)
    STEP 8  Build per-host compliance rows + review/FYI lists
    STEP 9  Write CIS Compliance Report.xlsx
    STEP 10 Write CIS Review.xlsx
    """
    print_banner()

    # ── STEP 1: Guide folder ──────────────────────────────────────────────────
    step(1, "Hardening Guide Folder")
    print(dim('  Enter the folder containing the Excel hardening guide templates.'))
    print(dim('  Tip: wrap paths with spaces in double quotes.\n'))
    guide_dir  = prompt_path("  Guide folder path: ")
    xlsx_files = sorted(f for f in os.listdir(guide_dir) if f.lower().endswith(".xlsx"))
    if not xlsx_files: err("No .xlsx files found in: "+guide_dir); sys.exit(1)
    print("\n  "+green(str(len(xlsx_files)))+" guide template(s) found:")
    for i, f in enumerate(xlsx_files, 1):
        gtype = detect_guide_type(f); tag = GUIDE_TAGS.get(gtype, yellow("[Unknown  ]"))
        print("    "+dim(str(i)+".").ljust(6)+" "+tag+" "+f)
    print()
    if not confirm("Proceed with these guides?"): print(yellow("  Aborted.")); sys.exit(0)

    # ── STEP 2: Scan folder ───────────────────────────────────────────────────
    step(2, "Nessus Scan Results Folder")
    print(dim("  Enter the root folder containing Nessus scan CSV file(s)."))
    print(dim("  Sub-folders are searched automatically.\n"))
    scan_dir = prompt_path("  Nessus scan folder path: ")
    csvs = sorted(os.path.join(root,f)
                  for root,_,files in os.walk(scan_dir)
                  for f in files if f.lower().endswith(".csv"))
    if not csvs: err("No CSV files found under: "+scan_dir); sys.exit(1)
    print("\n  "+green(str(len(csvs)))+" scan CSV file(s) found:")
    for f in csvs: found(os.path.relpath(f, scan_dir))
    print()
    if not confirm("Proceed with these scan files?"): print(yellow("  Aborted.")); sys.exit(0)

    # ── STEP 3: Output folder ─────────────────────────────────────────────────
    step(3, "Output Location")
    print(dim("  Press Enter to save to the scan folder, or enter a different path."))
    raw_out = strip_quotes(input("  Output folder ["+scan_dir+"]: "))
    out_dir = raw_out if raw_out and os.path.isdir(raw_out) else scan_dir
    ok("Reports will be saved to: "+out_dir)

    # ── STEP 4: Parse scan CSVs ───────────────────────────────────────────────
    step(4, "Loading Nessus Scan Files")
    all_scan_data = []
    for csv_path in csvs:
        print("\n  "+cyan(os.path.relpath(csv_path, scan_dir)))
        hosts, tidx = load_scan_per_host(csv_path)
        # Separate hosts with valid results from those where the audit failed to run
        valid   = {h:d for h,d in hosts.items() if host_has_results(d)}
        skipped = {h:d for h,d in hosts.items() if not host_has_results(d)}
        if valid:   ok("Valid hosts ("+str(len(valid))+"): "+", ".join(sorted(valid)))
        if skipped: warn("Skipped -- audit failed: "+", ".join(sorted(skipped)))
        if valid: all_scan_data.append((csv_path, valid, tidx))
    if not all_scan_data: err("No valid scan data found."); sys.exit(1)

    # ── STEP 5: Classify scans by OS ─────────────────────────────────────────
    step(5, "Classifying Scans")
    win_entries = []; rhel_entries = []
    for csv_path, valid, tidx in all_scan_data:
        fname = os.path.basename(csv_path).lower()
        gtype = detect_guide_type(fname)
        if gtype == "windows" or "windows" in fname:
            ver = detect_win_version(csv_path)
            win_entries.append({"csv_path":csv_path,"ver":ver or "?","valid":valid,"tidx":tidx})
            ok("Windows "+(ver or "?")+": "+os.path.basename(csv_path))
        elif any(k in fname for k in ("rhel","redhat","red hat","rehl","linux")):
            ver = detect_rhel_version(csv_path)
            cii = not any(k in fname for k in ("non","ncii","non-cii","non cii"))
            rhel_entries.append({"csv_path":csv_path,"ver":ver or "?","cii":cii,"valid":valid,"tidx":tidx})
            ok("RHEL"+(ver or "?")+" "+("CII" if cii else "non-CII")+": "+os.path.basename(csv_path))
        else:
            # Cannot auto-detect — ask the analyst
            warn("Cannot auto-detect type: "+os.path.basename(csv_path))
            print("    [1] Windows   [2] RHEL CII   [3] RHEL non-CII   [0] Skip")
            while True:
                choice = input("    Enter 0/1/2/3: ").strip()
                if choice == "0": break
                if choice == "1":
                    ver = detect_win_version(csv_path)
                    win_entries.append({"csv_path":csv_path,"ver":ver or "?","valid":valid,"tidx":tidx}); break
                if choice == "2":
                    ver = detect_rhel_version(csv_path)
                    rhel_entries.append({"csv_path":csv_path,"ver":ver or "?","cii":True,"valid":valid,"tidx":tidx}); break
                if choice == "3":
                    ver = detect_rhel_version(csv_path)
                    rhel_entries.append({"csv_path":csv_path,"ver":ver or "?","cii":False,"valid":valid,"tidx":tidx}); break

    # ── STEP 6: Guide matching ────────────────────────────────────────────────
    step(6, "Guide Matching")
    print(dim("  Each scan is matched to its hardening guide by OS version."))
    print(dim("  You can override any assignment before proceeding."))
    win_map, rhel_map = guide_matching_step(win_entries, rhel_entries, guide_dir, xlsx_files)

    # ── STEP 7: Load guides (cached) ──────────────────────────────────────────
    step(7, "Loading Guides")
    guide_cache = {}
    def get_guide(gf):
        """Load a guide from disk, or return the cached version if already loaded."""
        if gf not in guide_cache:
            print("  "+GUIDE_TAGS.get(detect_guide_type(gf), cyan("[Guide]  "))+" "+gf)
            guide_cache[gf] = load_guide_excel(os.path.join(guide_dir, gf))
        return guide_cache[gf]
    for gf in sorted(set(list(win_map.values())+list(rhel_map.values()))): get_guide(gf)
    ok("Guides loaded: "+", ".join(f+"("+str(len(guide_cache[f]))+")" for f in sorted(guide_cache)))

    # ── STEP 8: Build compliance rows ─────────────────────────────────────────
    step(8, "Building Compliance Sheets")
    all_sheets = []
    # Accumulators for cross-host review/FYI processing
    rhel_cii_valid={}; rhel_non_valid={}; rhel_cii_tidx={}; rhel_non_tidx={}
    rhel_cii_gf=None; rhel_non_gf=None

    # Windows hosts
    for e in win_entries:
        if e["csv_path"] not in win_map: continue
        gf = win_map[e["csv_path"]]; guide = get_guide(gf); ver = e["ver"]
        for host in sorted(e["valid"]):
            sname = "Win"+ver+"-"+host; label = "Windows Server "+ver+" MS-CII"
            rows = build_host_rows(guide, e["valid"][host], e["tidx"].get(host,{}))
            all_sheets.append((sname, host, label, rows, gf))
            p = sum(1 for r in rows if "PASSED" in str(r[4]))
            f = sum(1 for r in rows if "FAILED" in str(r[4]))
            print("  "+green("+")+" "+sname+"   "+green("PASS:"+str(p))+"  "+red("FAIL:"+str(f)))

    # RHEL hosts
    for e in rhel_entries:
        if e["csv_path"] not in rhel_map: continue
        gf = rhel_map[e["csv_path"]]; guide = get_guide(gf); ver = e["ver"]
        kind = "CII" if e["cii"] else "nCII"
        tidx_e = e.get("tidx", {})
        for host in sorted(e["valid"]):
            sname = "RHEL"+ver+"-"+kind+"-"+host; label = "RHEL"+ver+" "+kind
            rows = build_host_rows(guide, e["valid"][host], tidx_e.get(host,{}))
            all_sheets.append((sname, host, label, rows, gf))
            p = sum(1 for r in rows if "PASSED" in str(r[4]))
            f = sum(1 for r in rows if "FAILED" in str(r[4]))
            print("  "+green("+")+" "+sname+"   "+green("PASS:"+str(p))+"  "+red("FAIL:"+str(f)))
        if e["cii"]: rhel_cii_valid.update(e["valid"]); rhel_cii_tidx.update(tidx_e); rhel_cii_gf=gf
        else:        rhel_non_valid.update(e["valid"]); rhel_non_tidx.update(tidx_e); rhel_non_gf=gf

    # Build review and FYI lists across all hosts per guide
    print("\n  "+dim("Building Review and FYI lists..."))
    review = []; fyi = []
    win_groups = {}
    for e in win_entries:
        gf = win_map.get(e["csv_path"])
        if not gf: continue
        if gf not in win_groups: win_groups[gf] = {"valid":{},"tidx":{},"ver":e["ver"]}
        win_groups[gf]["valid"].update(e["valid"]); win_groups[gf]["tidx"].update(e["tidx"])
    for gf, grp in win_groups.items():
        g = get_guide(gf); lbl = "Windows "+grp["ver"]+" MS-CII"
        print("  "+dim("  Review: "+lbl+" ..."), end="", flush=True)
        rv = build_review(g, grp["valid"], lbl, grp["tidx"]); review += rv
        print(green(" "+str(len(rv))+" items"))
        print("  "+dim("  FYI:    "+lbl+" ..."), end="", flush=True)
        fv = build_fyi(grp["valid"], g, lbl, grp["tidx"]); fyi += fv
        print(green(" "+str(len(fv))+" items"))
    if rhel_cii_gf and rhel_cii_valid:
        g = get_guide(rhel_cii_gf)
        print("  "+dim("  Review: RHEL CII ..."), end="", flush=True)
        rv = build_review(g, rhel_cii_valid, "RHEL CII", rhel_cii_tidx); review += rv
        print(green(" "+str(len(rv))+" items"))
        print("  "+dim("  FYI:    RHEL CII ..."), end="", flush=True)
        fv = build_fyi(rhel_cii_valid, g, "RHEL CII", rhel_cii_tidx); fyi += fv
        print(green(" "+str(len(fv))+" items"))
    if rhel_non_gf and rhel_non_valid:
        g = get_guide(rhel_non_gf)
        print("  "+dim("  Review: RHEL non-CII ..."), end="", flush=True)
        rv = build_review(g, rhel_non_valid, "RHEL non-CII", rhel_non_tidx); review += rv
        print(green(" "+str(len(rv))+" items"))
        print("  "+dim("  FYI:    RHEL non-CII ..."), end="", flush=True)
        fv = build_fyi(rhel_non_valid, g, "RHEL non-CII", rhel_non_tidx); fyi += fv
        print(green(" "+str(len(fv))+" items"))
    ok("Review: "+str(len(review))+" items  |  FYI: "+str(len(fyi))+" items")

    # ── STEP 9: Write CIS Compliance Report.xlsx ──────────────────────────────
    step(9, "Writing CIS Compliance Report.xlsx")
    wb = openpyxl.Workbook(); ws_s = wb.active; ws_s.title = "Summary"

    # Summary sheet header
    ws_s["A1"] = "CIS Hardening Compliance Report - Summary"
    ws_s["A1"].font = FNT_TITLE
    ws_s.merge_cells("A1:"+get_column_letter(9)+"1")
    ws_s["A1"].alignment = Alignment(horizontal="center")
    ws_s.append([])

    # Summary table — one row per host sheet
    hdrs_s = ["Sheet Name","Host IP","Type","PASSED","FAILED",
              "NOT IN SCAN","Not Adopted\n(Intentional)","N/A","Total"]
    ws_s.append(hdrs_s); style_hdr(ws_s, 3, len(hdrs_s))
    for sname, host, label, rows, gf in all_sheets:
        p  = sum(1 for r in rows if "PASSED"      in str(r[4]))
        f  = sum(1 for r in rows if "FAILED"      in str(r[4]))
        ni = sum(1 for r in rows if r[4] == "NOT IN SCAN")
        no = sum(1 for r in rows if "NOT ADOPTED" in str(r[4]).upper())
        na = sum(1 for r in rows if str(r[4]).strip() == "N/A")
        ws_s.append([sname, host, label, p, f, ni, no, na, len(rows)])
        rn = ws_s.max_row
        for c in range(1, len(hdrs_s)+1):
            ws_s.cell(rn,c).border = BORDER; ws_s.cell(rn,c).alignment = ALIGN_CTR
        ws_s.cell(rn,4).fill=FILL_PASS; ws_s.cell(rn,5).fill=FILL_FAIL
        ws_s.cell(rn,6).fill=FILL_WARN; ws_s.cell(rn,7).fill=FILL_NO; ws_s.cell(rn,8).fill=FILL_NA

    # Legend notes at the bottom of the summary sheet
    ws_s.append([])
    for note in [
        '"NOT IN SCAN" = guide item (Adoption=YES) with no matching Nessus check. See Review workbook.',
        '"guide X -> scan Y" = CIS S/N renumbered between guide and Nessus audit; result is valid.',
        '"Not Adopted" = intentionally excluded per hardening guide (business justification exists).',
        '"N/A" = not applicable to this server role.',
    ]: ws_s.append([note])

    set_widths(ws_s, [28,14,24,10,10,14,16,10,10])

    # Write one sheet per host
    for sname, host, label, rows, _ in all_sheets:
        write_host_sheet(wb, sname, rows, host, label)

    out_main = os.path.join(out_dir, "CIS Compliance Report.xlsx")
    wb.save(out_main); ok("Saved: "+out_main)

    # ── STEP 10: Write CIS Review.xlsx ────────────────────────────────────────
    step(10, "Writing CIS Review.xlsx")
    wb_r = openpyxl.Workbook()
    RH = ["Source","Host(s)","Guide S/N","Title","Adoption","Justification","Reason","Detail / Notes"]
    RW = [26,16,16,50,11,34,42,60]

    # Summary sheet
    ws_rs = wb_r.active; ws_rs.title = "Summary"
    ws_rs["A1"] = "CIS Hardening - Review Items"
    ws_rs["A1"].font = Font(bold=True, size=13, color="843C0C")
    ws_rs.merge_cells("A1:B1"); ws_rs["A1"].alignment = Alignment(horizontal="center")
    ws_rs.append([])
    ws_rs.append(["Manual review items (guide checks not found in Nessus scan):", len(review)])
    ws_rs.append(["FYI items (Nessus checks not in hardening guide - newer CIS version):", len(fyi)])
    ws_rs.append([])
    ws_rs.append(["By Source","Count"])
    style_hdr(ws_rs, ws_rs.max_row, 2, PatternFill("solid", fgColor="843C0C"))
    for src, cnt in sorted(Counter(r[0] for r in review).items()): ws_rs.append([src, cnt])
    ws_rs.append([])
    ws_rs.append(["Action Guidance",""])
    style_hdr(ws_rs, ws_rs.max_row, 2, PatternFill("solid", fgColor="595959"))
    for note in [
        '"Not found in Nessus scan (any host)" -> Verify manually on all servers.',
        '"Not found in Nessus scan (X/Y hosts)" -> Verify on the listed hosts manually.',
        "FYI sheet: newer CIS benchmark checks, not covered by the current hardening guide.",
    ]: ws_rs.append(["- "+note,""])
    set_widths(ws_rs, [70,10])

    # Review Items sheet — controls needing manual verification
    ws_rv = wb_r.create_sheet("Review Items")
    ws_rv.append(RH); style_hdr(ws_rv, 1, len(RH), PatternFill("solid", fgColor="843C0C"))
    for row in review:
        ws_rv.append(list(row)); rn = ws_rv.max_row
        for c in range(1, len(RH)+1):
            ws_rv.cell(rn,c).border=BORDER; ws_rv.cell(rn,c).fill=FILL_REV
            ws_rv.cell(rn,c).alignment=ALIGN_CTR if c in (1,3,5) else ALIGN_WR
    set_widths(ws_rv, RW); ws_rv.freeze_panes="A2"; ws_rv.auto_filter.ref=ws_rv.dimensions

    # FYI sheet — Nessus checks not in the current hardening guide version
    ws_fy = wb_r.create_sheet("Newer CIS Version (FYI)")
    ws_fy.append(RH); style_hdr(ws_fy, 1, len(RH), PatternFill("solid", fgColor="595959"))
    for row in fyi:
        ws_fy.append(list(row)); rn = ws_fy.max_row
        for c in range(1, len(RH)+1):
            ws_fy.cell(rn,c).border=BORDER; ws_fy.cell(rn,c).fill=FILL_GRAY
            ws_fy.cell(rn,c).alignment=ALIGN_CTR if c in (1,3,5) else ALIGN_WR
    set_widths(ws_fy, RW); ws_fy.freeze_panes="A2"; ws_fy.auto_filter.ref=ws_fy.dimensions

    out_rev = os.path.join(out_dir, "CIS Review.xlsx")
    wb_r.save(out_rev); ok("Saved: "+out_rev)

    # ── Completion summary ────────────────────────────────────────────────────
    sep = "="*62
    print(); print(cyan(sep))
    print(cyan("  ")+bold(green("  "+TOOL_NAME+"  -  Completed successfully")))
    print(cyan("  ")+"  Host sheets generated : "+bold(str(len(all_sheets))))
    print(cyan("  ")+"  Review items          : "+bold(yellow(str(len(review)))))
    print(cyan("  ")+"  FYI items             : "+bold(dim(str(len(fyi)))))
    print(cyan("  "))
    print(cyan("  ")+"  "+bold("Output files:"))
    print(cyan("  ")+"    "+green(out_main))
    print(cyan("  ")+"    "+green(out_rev))
    print(cyan(sep)); print()

if __name__ == "__main__":
    main()
